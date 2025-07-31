import openpyxl
import re
import requests
import os

import pandas as pd
from difflib import get_close_matches
import unicodedata
import re
from django.shortcuts import redirect
from django.shortcuts import HttpResponse
import threading

from datetime import datetime
from unidecode import unidecode

from django.conf import settings

from invest.models import *
from invest.utils import NormalizarCadena
import json


def ImportarActivos(campanya, usuario, actualizar=False):
    # Mapeo de columnas por proveedor/cartera
    ruta_json = os.path.join(settings.BASE_DIR, "invest", "MAPEO_COLUMNAS.json")

    with open(ruta_json, "r", encoding="utf-8") as f:
        MAPEO_COLUMNAS = json.load(f)
    proveedor_o_cartera = campanya.proveedor.nombre.upper()
    mapeo_columnas = MAPEO_COLUMNAS.get(proveedor_o_cartera, {})
    wb_obj = openpyxl.load_workbook(filename=campanya.fichero)
    sheet_obj = wb_obj.worksheets[0]
    num_columnas = sheet_obj.max_column
    lineas_vacias = 0
    encontradotexto = False
    contador_vacios = 0
    if actualizar:
        id_activos = CampanyaLinea.objects.filter(campanya=campanya).values_list(
            "activo__pk", flat=True
        )
    primero = True

    for i in range(1, 30000):
        tipologia = ""
        if lineas_vacias > 10:
            break
        if not encontradotexto:
            v_columnas = []
            for columna in range(0, num_columnas + 1):
                try:
                    celda = (
                        str(sheet_obj.cell(row=i, column=columna).value).strip().upper()
                    )
                except:
                    celda = ""
                # Buscar columnas según el mapeo
                for etiqueta, nombres in mapeo_columnas.items():
                    if celda in nombres:
                        v_columnas.append(etiqueta)
                        encontradotexto = True
                        break
                else:
                    v_columnas.append("")

                if v_columnas == ["", "", "", "", "", "", "", "", ""]:
                    break
        else:
            lineas_vacias += 1
            if settings.DEBUG and primero:
                primero = False

            id_proveedor = ""
            ref_ue = ""
            tipo = ""
            ref_catastral = ""
            cp = ""
            direccion = ""
            municipio_nombre = ""
            provincia_nombre = ""
            asset_id = ""
            m2 = None
            fecha_construccion = None
            num_habitaciones = None
            num_banyos = None
            estado = ""
            precio = 0
            estado_ocupacional = "0"
            estado_legal = ""
            url = ""
            alquilado = False
            observaciones = ""
            longitud = ""
            latitud = ""
            importe_deuda = None
            valor_referencia = None
            clase = campanya.tipo
            campo_id = ""

            for columna in range(len(v_columnas)):

                if contador_vacios == 10:
                    break
                if v_columnas[columna] != "":
                    if (
                        "=VLOOKUP" in celda or "=BUSCARV" in celda
                    ):  # quito los campos calculados
                        celda = ""
                    try:
                        celda = str(sheet_obj.cell(row=i, column=columna).value)
                    except:
                        celda = ""
                    if celda.lower() == "none":
                        celda = ""
                    if celda in ["NPL", "CDR", "REO"]:
                        v_columnas[columna] = "clase"
                    if v_columnas[columna] == "ID":
                        print(celda)
                        if celda.strip() != "" and not id_proveedor:
                            id_proveedor = celda.replace(",", ";")
                            print(id_proveedor)
                    elif v_columnas[columna] == "RefUE":
                        ref_ue = celda
                    elif v_columnas[columna] == "Observaciones":
                        if observaciones == "":
                            observaciones = celda
                        else:
                            if celda != "":
                                observaciones += " / " + celda
                    elif v_columnas[columna] == "ASSETID":
                        asset_id = celda
                    elif v_columnas[columna] == "Tipo":
                        tipo_ = celda
                        if not actualizar:
                            tipologia_ = ComprobarTipologia(tipo_, usuario)
                            if tipologia_:
                                tipo = tipo_
                            else:
                                if not tipo:
                                    tipo = tipo_
                    elif v_columnas[columna] == "Ref. Catastral":
                        if ref_catastral:
                            if len(celda) == 20:
                                if ";" in celda:
                                    ref_catastral = celda.replace(";", " ")
                                else:
                                    ref_catastral = celda

                        else:
                            if ";" in celda:
                                ref_catastral = celda.replace(";", " ")
                            else:
                                ref_catastral = celda
                    elif v_columnas[columna] == "CP":
                        cp = celda
                    elif v_columnas[columna] == "Dirección":
                        if direccion:
                            direccion += f" {celda}"
                        else:
                            direccion = celda
                    elif v_columnas[columna] == "Municipio":
                        municipio_nombre = normalizar_municipio(celda)
                    elif v_columnas[columna] == "Provincia":
                        provincia_nombre = celda
                    elif v_columnas[columna] == "Estado":
                        estado = celda
                    elif v_columnas[columna] == "m2":
                        m2 = celda.replace(".", ",")
                    elif v_columnas[columna] == "Precio":
                        celda_limpia = celda
                        celda_limpia = str(celda_limpia).replace(" ", "")
                        celda_limpia = str(celda_limpia).replace("€", "")
                        if celda_limpia.count(".") > 0 and celda_limpia.count(",") > 0:
                            celda_limpia = celda_limpia.replace(".", "")
                            celda_limpia = str(celda_limpia).replace(",", ".")
                        try:
                            precio = round(float(celda_limpia))
                        except:
                            precio = None

                    elif v_columnas[columna] == "Importe Deuda":
                        try:
                            #                            celda_limpia = str(celda).replace('.', '')
                            celda_limpia = celda
                            celda_limpia = str(celda_limpia).replace(" ", "")
                            celda_limpia = str(celda_limpia).replace("€", "")
                            if (
                                celda_limpia.count(".") > 0
                                and celda_limpia.count(",") > 0
                            ):

                                celda_limpia = celda_limpia.replace(".", "")
                                celda_limpia = str(celda_limpia).replace(",", ".")
                            celda_valor = float(celda_limpia)

                            if (
                                importe_deuda is None
                                or celda_valor > importe_deuda
                                or celda_valor == 0
                            ):
                                importe_deuda = celda_valor
                        except (ValueError, TypeError) as e:
                            print(f"Error al convertir: {e}")
                            pass
                    elif v_columnas[columna] == "valor_referencia":
                        try:
                            valor_referencia = float(celda)
                        except:
                            valor_referencia = None
                    elif v_columnas[columna] == "estado_ocupacional":
                        if celda == "1. Empty":
                            estado_ocupacional = "1"
                        elif celda == "2. Rented":
                            estado_ocupacional = "2"
                        elif celda == "3. Squatted":
                            estado_ocupacional = "3"
                        elif celda == "4. Occupied by owner":
                            estado_ocupacional = "4"
                    elif v_columnas[columna] in [
                        "Fase legal 1",
                        "Situación Saneamiento",
                        "Estado saneamiento",
                    ]:
                        estado_legal += f" {celda}"
                    elif v_columnas[columna] == "Fase legal 2":
                        estado_legal += f" {celda}"
                    elif v_columnas[columna] == "Alquilado":
                        if celda == "Yes" or celda == "Sí":
                            estado_ocupacional = "2"
                    elif v_columnas[columna] == "URL":
                        url = celda
                    elif v_columnas[columna] == "Longitud":
                        longitud = celda
                    elif v_columnas[columna] == "Latitud":
                        latitud = celda
                    elif v_columnas[columna] == "CAMPOID":
                        campo_id = celda
                    elif v_columnas[columna] == "Clase":
                        if celda:
                            if celda.lower() == "venta de deuda":
                                clase = "NPL"
                            elif celda.lower() == "venta de colateral":
                                clase = "REO"
                            else:
                                if celda.upper() in ["NPL", "CDR", "REO"]:
                                    clase = celda
                                else:
                                    print(celda)

            if id_proveedor == "None":
                id_proveedor = ""
            if not id_proveedor and campo_id:
                id_proveedor = campo_id

            if (
                (ref_ue and ref_ue != "None")
                or (id_proveedor and id_proveedor != "None")
                or (ref_catastral and ref_catastral != "None")
                or (direccion and direccion != "None")
            ):
                tipologia_ = ComprobarTipologia(tipo, usuario)
                if tipologia_:
                    tipologia = tipologia_

                # municipio = Poblacion.objects.filter(nombre__iexact=NormalizarCadena(municipio_nombre)).first()
                municipio = Poblacion.objects.filter(
                    nombre__iexact=municipio_nombre
                ).first()

                if not municipio:
                    municipios = municipio_nombre.split("/")
                    for municipio_nombre in municipios:
                        # municipio = Poblacion.objects.filter(nombre__iexact=NormalizarCadena(municipio_nombre))
                        municipio = Poblacion.objects.filter(
                            nombre__iexact=municipio_nombre
                        )

                        if municipio.first():
                            if municipio.count() > 1:
                                provincia = DameProvincia(provincia_nombre)
                                municipio = municipio.filter(
                                    provincia=provincia
                                ).first()
                            else:
                                municipio = municipio.first()
                            break

                if not municipio:
                    if provincia_nombre.lower() == "lérida":
                        provincia_nombre = "Lleida"
                    if provincia_nombre.lower() == "gerona":
                        provincia_nombre = "Girona"
                    provincia = DameProvincia(provincia_nombre)
                    municipio = Poblacion.create(provincia, municipio_nombre)
                try:
                    m2 = float(m2) if m2 not in (None, "") else 0
                except:
                    m2 = 0

                if actualizar and id_proveedor:

                    if (
                        ref_catastral
                        and len(ref_catastral) > 8
                        and not " " in ref_catastral
                    ):
                        activo = Activo.objects.filter(
                            ref_catastral=ref_catastral, pk__in=id_activos
                        ).first()
                        activo.id_proveedor = id_proveedor
                        activo.save()
                        lineas_vacias = 0
                else:

                    if not ref_catastral:
                        activo = None
                    else:
                        ref_catastral = ref_catastral.strip()

                    if (
                        ref_catastral
                        and len(ref_catastral) > 8
                        and not " " in ref_catastral
                    ):
                        if id_proveedor:
                            activo = Activo.objects.filter(
                                ref_catastral=ref_catastral, id_proveedor=id_proveedor
                            ).first()
                        else:
                            activo = None
                    else:
                        activo = None
                    if ref_ue == "None":
                        ref_ue = ""
                    direccion = LimpiaDireccion(direccion)
                    activos = []
                    filtro = {"id_proveedor": id_proveedor}
                    if ref_catastral:
                        filtro["ref_catastral"] = ref_catastral
                    if direccion:
                        filtro["direccion"] = direccion

                    activo = Activo.objects.filter(**filtro).first()
                    if not activo:

                        activo = Activo.objects.create(
                            id_proveedor=id_proveedor[:200],
                            ref_ue=ref_ue[:200],
                            tipologia=tipologia,
                            ref_catastral=ref_catastral[:200],
                            cp=cp[:20],
                            direccion=direccion,
                            poblacion=municipio,
                            m2=m2,
                            fecha_construccion=fecha_construccion,
                            num_habitaciones=num_habitaciones,
                            num_banyos=num_banyos,
                        )
                        if longitud:
                            try:
                                activo.longitud = float(longitud)
                                activo.save()
                            except (ValueError, TypeError):
                                pass
                        if latitud:
                            try:
                                activo.latitud = float(latitud)
                                activo.save()
                            except (ValueError, TypeError):
                                pass
                        activos.append(activo)
                        lineas_vacias = 0
                    else:
                        # controlar duplicados
                        actualizar_activo = False

                        es_diferente = False
                        if activo.id_proveedor != id_proveedor:
                            es_diferente = True
                        elif (
                            ref_catastral
                            and activo.ref_catastral
                            and activo.ref_catastral != ref_catastral
                        ):
                            es_diferente = True
                        elif (
                            direccion
                            and activo.direccion
                            and activo.direccion != direccion
                        ):
                            es_diferente = True
                        if es_diferente:
                            activo.id_proveedor = id_proveedor
                            if ref_ue is not None:
                                activo.ref_ue = ref_ue
                            if tipologia is not None:
                                activo.tipologia = tipologia
                            if ref_catastral is not None:
                                activo.ref_catastral = ref_catastral
                            if cp is not None:
                                activo.cp = cp
                            if direccion is not None:
                                activo.direccion = direccion
                            if municipio is not None:
                                activo.poblacion = municipio
                            if m2 is not None:
                                activo.m2 = m2
                            if fecha_construccion is not None:
                                activo.fecha_construccion = fecha_construccion
                            if num_habitaciones is not None:
                                activo.num_habitaciones = num_habitaciones
                            if num_banyos is not None:
                                activo.num_banyos = num_banyos
                            activo.save()

                        lineas_vacias = 0
                        if longitud:
                            try:
                                activo.longitud = float(longitud)
                                activo.save()
                            except (ValueError, TypeError):
                                pass
                        if latitud:
                            try:
                                activo.latitud = float(latitud)
                                activo.save()
                            except (ValueError, TypeError):
                                pass
                        activos.append(activo)
                        lineas_vacias = 0
                    if activos:
                        for (
                            activo
                        ) in (
                            activos
                        ):  # por si hay varias referencias catastrales en una línea
                            CampanyaLinea.objects.create(
                                campanya=campanya,
                                tipo=clase,
                                activo=activo,
                                precio=precio,
                                importe_deuda=importe_deuda,
                                valor_referencia=valor_referencia,
                                estado_ocupacional=estado_ocupacional,
                                estado_legal=estado_legal,
                                url=url,
                                observaciones=observaciones,
                            )
                            if settings.DEBUG:
                                print(
                                    f"{i}: {activo.ref_ue} {activo.ref_catastral} {activo.direccion}"
                                )
                                # if i == 400:
                                # breakpoint()
                                if (
                                    activo.ref_ue
                                    or activo.ref_catastral
                                    or activo.direccion
                                ):
                                    lineas_vacias = 0
            # else:
            # if settings.DEBUG:
            # print(str(lineas_vacias))

    id_activos = CampanyaLinea.objects.filter(campanya=campanya).values_list(
        "activo__pk", flat=True
    )
    activos = Activo.objects.filter(pk__in=id_activos)
    # activos2 = Activo.objects.filter(fecha_peticion_catastro__isnull=False, catastro=False)

    # for activo in activos:
    # activo.ActivoLocalizar()
    threading.Thread(
        target=lambda: [activo.ActivoLocalizar() for activo in activos]
    ).start()
    # ver todos los registros que no se ha podido hacer la peticion por sobrecarga de peticiones
    # threading.Thread(target=lambda: [activo.ActivoLocalizar() for activo in activos2]).start()


def ComprobarTipologia(tipo, usuario, crearsinoexiste=True):

    tipo = tipo.upper()
    grupo = None
    subgrupo = None
    if "VIVIENDA" in tipo or "PISO" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="VIVIENDA").first()
    elif (
        "SUELO" in tipo
        or "DESARROLLO" in tipo
        or "TERCIARIO" in tipo
        or "RUSTICO" in tipo
        or "URBANO" in tipo
        or "SOLAR" in tipo
        or "CHALET" in tipo
        or "URBANIZABLE" in tipo
        or "RESIDENCIAL" in tipo
        or "LAND" in tipo
        or "FLAT" in tipo
        or "HOUSE" in tipo
    ):
        if "TERCIARIO" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(
                nombre="SUELO TERCIARIO"
            ).first()
        elif "DESARROLLO" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(
                nombre="DESARROLLO URBANÍSTICO"
            ).first()
        elif "RUSTICO" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="SUELO RUSTICO").first()
        elif "URBANO" in tipo or "URBANIZABLE" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="SUELO URBANO").first()
        else:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="SUELO").first()

    elif "GARAJE" in tipo or "PARKING" in tipo or "TRASTERO":
        subgrupo = SubGrupoTipologia.objects.filter(nombre="GARAJE").first()

    elif "LOCAL" in tipo or "RETAIL" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="LOCAL").first()
    elif (
        "NAVE" in tipo
        or "INDUSTRIAL" in tipo
        or "WAREHOUSE" in tipo
        or "STORAGE" in tipo
    ):
        subgrupo = SubGrupoTipologia.objects.filter(nombre="NAVE").first()
    elif "EDIFICIO" in tipo or "WIP" in tipo or "FINCA" in tipo:
        if "WIPS" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="EDIFICIO WIPS").first()
        else:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="EDIFICIO").first()
    elif (
        "ALOJAMIENTO" in tipo
        or "HOTEL" in tipo
        or "ALBERGUE" in tipo
        or "HOSTAL" in tipo
        or "CAMPING" in tipo
    ):
        if "ALBERGUE" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="ALBERGUE").first()
        elif "HOSTAL" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="HOSTAL").first()
        elif "PENSION" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="PENSION").first()
        elif "CAMPING" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="CAMPING").first()
        elif "HOTEL" in tipo:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="HOTEL").first()
        else:
            subgrupo = SubGrupoTipologia.objects.filter(nombre="ALOJAMIENTO").first()

    elif "CENTRO COMERCIAL" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="CENTRO COMERCIAL").first()
    elif "GERIÁTRICO" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="GERIÁTRICO").first()
    elif "GASOLINERA" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="GASOLINERA").first()
    elif "FARMACIA" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="FARMACIA").first()
    elif "AMARRE" in tipo or "PUERTO" in tipo:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="AMARRE-PUERTO").first()
    else:
        subgrupo = SubGrupoTipologia.objects.filter(nombre="VARIOS").first()

    grupo = GrupoTipologia.objects.filter(id=subgrupo.grupo_id).first()
    tipologia = Tipologia.objects.filter(
        nombre=NormalizarCadena(tipo), anulado_por=None
    ).first()
    print(tipologia)
    if not tipologia:
        tipologia = Tipologia.create(NormalizarCadena(tipo), usuario, grupo, subgrupo)

    return tipologia


def DameProvincia(provincia_nombre):
    provincia_nombre = provincia_nombre.capitalize()
    provincia = Provincia.objects.filter(nombre__iexact=provincia_nombre).first()
    if not provincia:
        provincia_partes = re.split(r"[ /]", provincia_nombre)
        qs_provincias = None
        for provincia_parte in provincia_partes:
            if len(provincia_partes) > 1:
                prov = DameProvincia(provincia_parte)
                if prov.pk != 0:
                    return prov
            if qs_provincias:
                try:
                    provincia_parte = int(provincia_parte)
                    qs_provincias2 = Provincia.objects.filter(codigo=provincia_parte)
                except:
                    qs_provincias2 = Provincia.objects.filter(
                        nombre__in=provincia_parte
                    )
                if qs_provincias2:
                    qs_provincias = qs_provincias2
            else:
                try:
                    provincia_parte = int(provincia_parte)
                    qs_provincias = Provincia.objects.filter(codigo=provincia_parte)
                except:
                    qs_provincias = Provincia.objects.filter(nombre__in=provincia_parte)
        if qs_provincias:
            provincia = qs_provincias.first()
        else:
            provincia = Provincia.objects.filter(nombre="Varios").first()
            if not provincia:
                if settings.DEBUG:
                    breakpoint()
                provincia = Provincia.create(0, "Varios")
    return provincia


def LimpiaDireccion(cadena):
    cadena.replace("Tipo vía: ", "")
    cadena.replace("; Nombre vía:", "")
    cadena.replace("; Bloque:    ", "")
    cadena.replace("; Escalera:   ", ", ")
    cadena.replace("; PLanta:     ", ", ")
    cadena.replace("; Puerta:     ", ", ")
    return cadena


def cargar_municipios():
    try:
        ruta_csv = os.path.join("fich_importaciones", "pueblos.csv")
        df = pd.read_csv(
            ruta_csv,
            header=None,
            names=[
                "cod_provincia",
                "provincia",
                "municipio",
                "cp",
                "latitud",
                "longitud",
            ],
        )  # quitar apostrofes l'
        df["municipio"] = df["municipio"].str.replace(r"l'", "", regex=True).str.strip()

        municipios = set(df["municipio"].str.strip().dropna().unique())

        municipio_map = {}
        for m in df["municipio"]:
            if pd.isna(m):
                continue
            m = m.strip()
            normalized = unidecode(m).lower().strip()
            municipio_map[normalized] = m

        return municipios, municipio_map
    except Exception as e:
        print(f"Error cargando municipios: {e}")
        return set(), {}


MUNICIPIOS_CORRECTOS, MUNICIPIO_MAP = cargar_municipios()


def normalizar_municipio(nombre):
    if not nombre or pd.isna(nombre):
        return ""

    nombre = str(nombre).strip()
    if nombre in MUNICIPIOS_CORRECTOS:
        return nombre

    def limpiar_texto(texto):
        texto.replace(",", "")
        texto = re.sub(r"^[lL]['’]\s*", "", texto)

        texto = unidecode(texto)
        texto = re.sub(r"[^a-zA-Z0-9 ñÑáéíóúÁÉÍÓÚàèòÀÈÒ]", "", texto)
        return texto.strip().lower()

    nombre_limpio = limpiar_texto(nombre)

    for normalized, original in MUNICIPIO_MAP.items():
        if normalized == nombre_limpio:
            return original

    coincidencias = get_close_matches(nombre, MUNICIPIOS_CORRECTOS, n=1, cutoff=0.6)
    if coincidencias:
        return coincidencias[0]

    return nombre
